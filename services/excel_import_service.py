"""
Excel Import Service - Framework-agnostic business logic.

This module contains the core Excel import functionality extracted from
the CLI tool, with progress callback support for API integration.
"""

import os
import hashlib
import logging
import json
import subprocess
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any, Set, Callable
from decimal import Decimal
import re

from sqlalchemy.orm import Session
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import networkx as nx

from backend.models.schema import Model, Cell
from services.formula_service import FormulaParser

logger = logging.getLogger(__name__)

# Default configuration (can be overridden)
DEFAULT_TOLERANCE = 1e-6
DEFAULT_MAX_CIRCULAR_ITERATIONS = 100
DEFAULT_CONVERGENCE_THRESHOLD = 1e-6
DEFAULT_MODELS_DIR = 'models/'
DEFAULT_HYPERFORMULA_WRAPPER = 'scripts/hyperformula_wrapper.js'


class CircularReferenceDetector:
    """Detect and analyze circular references in formulas."""
    
    def __init__(self):
        self.graph = nx.DiGraph()
        self.circular_groups: List[List[str]] = []
    
    def add_dependency(self, cell: str, depends_on: List[str]):
        """Add a cell and its dependencies to the graph."""
        self.graph.add_node(cell)
        for dep in depends_on:
            self.graph.add_edge(cell, dep)
    
    def detect_cycles(self) -> List[List[str]]:
        """
        Detect all circular reference groups.
        
        Returns list of circular reference groups (strongly connected components).
        """
        try:
            # Find strongly connected components (cycles)
            cycles = list(nx.strongly_connected_components(self.graph))
            # Filter out single nodes (not cycles)
            self.circular_groups = [list(cycle) for cycle in cycles if len(cycle) > 1]
            
            logger.info(f"Detected {len(self.circular_groups)} circular reference groups")
            for i, group in enumerate(self.circular_groups):
                logger.debug(f"Circular group {i+1}: {group}")
            
            return self.circular_groups
        except Exception as e:
            logger.error(f"Error detecting cycles: {e}")
            return []
    
    def is_circular(self, cell: str) -> bool:
        """Check if a cell is part of a circular reference."""
        for group in self.circular_groups:
            if cell in group:
                return True
        return False


class CircularSolver:
    """Iterative solver for circular references."""
    
    def __init__(self, max_iterations: int = DEFAULT_MAX_CIRCULAR_ITERATIONS, 
                 threshold: float = DEFAULT_CONVERGENCE_THRESHOLD):
        self.max_iterations = max_iterations
        self.threshold = threshold
    
    def solve(self, circular_cells: List[str], cell_data: Dict[str, Dict], 
              evaluate_func: Callable) -> Tuple[Dict[str, Any], str, int]:
        """
        Iteratively solve circular references.
        
        CRITICAL: NEVER copies raw_value to calculated_value. Sets NULL on failure.
        
        Returns: (results_dict, status, iterations)
            status: 'converged', 'max_iterations', or 'error'
        """
        logger.info(f"Starting iterative solver for {len(circular_cells)} circular cells")
        
        # Initialize with zeros for numeric, empty string for text
        values = {}
        for cell_ref in circular_cells:
            cell = cell_data.get(cell_ref)
            if cell and FormulaParser.is_text_formula(cell.get('formula', '')):
                values[cell_ref] = ''
            else:
                values[cell_ref] = 0.0
        
        converged_cells = set()
        
        for iteration in range(self.max_iterations):
            new_values = {}
            max_change = 0.0
            
            for cell_ref in circular_cells:
                # Note: Don't skip "converged" cells - in circular references,
                # all cells must continue evaluating together even if individually stable
                
                try:
                    # Evaluate with current context
                    result = evaluate_func(cell_ref, values)
                    
                    if result is None:
                        # Failed to evaluate - set NULL, DO NOT copy raw_value
                        logger.error(f"Failed to evaluate circular cell {cell_ref} "
                                   f"(formula: {cell_data.get(cell_ref, {}).get('formula', 'N/A')})")
                        new_values[cell_ref] = None
                        continue
                    
                    new_values[cell_ref] = result
                    
                    # Calculate change
                    if isinstance(result, (int, float)) and isinstance(values[cell_ref], (int, float)):
                        change = abs(result - values[cell_ref])
                        max_change = max(max_change, change)
                        
                        # Track converged cells for logging, but don't skip them
                        if change < self.threshold:
                            if cell_ref not in converged_cells:
                                converged_cells.add(cell_ref)
                                logger.debug(f"Cell {cell_ref} converged (change: {change:.2e})")
                        else:
                            # Cell started changing again, remove from converged set
                            converged_cells.discard(cell_ref)
                    
                except Exception as e:
                    logger.error(f"Error evaluating circular cell {cell_ref}: {e}")
                    new_values[cell_ref] = None
            
            values = new_values
            
            logger.debug(f"Iteration {iteration + 1}: max_change={max_change:.2e}, "
                        f"converged={len(converged_cells)}/{len(circular_cells)}")
            
            # Check global convergence based on max change across all cells
            if max_change < self.threshold:
                logger.info(f"Converged after {iteration + 1} iterations")
                return values, 'converged', iteration + 1
        
        logger.warning(f"Max iterations ({self.max_iterations}) reached without full convergence")
        return values, 'max_iterations', self.max_iterations


class HyperFormulaEvaluator:
    """Interface to HyperFormula via Node.js subprocess."""
    
    def __init__(self, wrapper_path: str = DEFAULT_HYPERFORMULA_WRAPPER):
        self.wrapper_path = wrapper_path
        
        if not Path(wrapper_path).exists():
            logger.warning(f"HyperFormula wrapper not found at {wrapper_path}")
    
    def evaluate_batch(self, sheets_data: List[Dict], queries: List[Dict]) -> Dict:
        """
        Evaluate multiple formulas using HyperFormula.
        
        Args:
            sheets_data: List of sheet definitions with cells
            queries: List of cells to evaluate
        
        Returns:
            Result dictionary with success flag and results
        """
        request = {
            'sheets': sheets_data,
            'queries': queries
        }
        
        try:
            # Call Node.js wrapper
            process = subprocess.Popen(
                ['node', self.wrapper_path],
                stdin=subprocess.PIPE,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )
            
            stdout, stderr = process.communicate(json.dumps(request), timeout=30)
            
            if process.returncode != 0:
                logger.error(f"HyperFormula error (exit {process.returncode}): {stderr}")
                return {'success': False, 'error': stderr}
            
            result = json.loads(stdout)
            logger.debug(f"HyperFormula evaluated {len(queries)} queries")
            return result
            
        except subprocess.TimeoutExpired:
            logger.error("HyperFormula evaluation timed out")
            return {'success': False, 'error': 'Timeout'}
        except Exception as e:
            logger.error(f"HyperFormula evaluation failed: {e}")
            return {'success': False, 'error': str(e)}


class ExcelImportService:
    """
    Framework-agnostic Excel import service.
    
    This service handles the complete Excel import workflow with progress tracking.
    """
    
    def __init__(
        self,
        db_session: Session,
        progress_callback: Optional[Callable[[str, float, str], None]] = None,
        tolerance: float = DEFAULT_TOLERANCE,
        max_circular_iterations: int = DEFAULT_MAX_CIRCULAR_ITERATIONS,
        convergence_threshold: float = DEFAULT_CONVERGENCE_THRESHOLD,
        models_dir: str = DEFAULT_MODELS_DIR,
        hyperformula_wrapper: str = DEFAULT_HYPERFORMULA_WRAPPER
    ):
        """
        Initialize Excel import service.
        
        Args:
            db_session: SQLAlchemy database session
            progress_callback: Optional callback for progress updates
                              Signature: callback(stage: str, percent: float, message: str)
            tolerance: Tolerance for mismatch detection (default: 1e-6)
            max_circular_iterations: Max iterations for circular solver
            convergence_threshold: Convergence threshold for circular solver
            models_dir: Directory to store Excel files
            hyperformula_wrapper: Path to HyperFormula wrapper script
        """
        self.session = db_session
        self.progress_callback = progress_callback or (lambda *args: None)
        
        self.tolerance = tolerance
        self.max_circular_iterations = max_circular_iterations
        self.convergence_threshold = convergence_threshold
        self.models_dir = models_dir
        self.hyperformula_wrapper = hyperformula_wrapper
        
        # Initialize components
        self.hf_evaluator = HyperFormulaEvaluator(hyperformula_wrapper)
        self.parser = FormulaParser()
        self.circular_detector = CircularReferenceDetector()
        self.circular_solver = CircularSolver(max_circular_iterations, convergence_threshold)
        
        # Statistics tracking
        self.stats = {
            'total_cells': 0,
            'value_cells': 0,
            'formula_cells': 0,
            'formula_text_cells': 0,
            'circular_references': 0,
            'circular_converged': 0,
            'circular_failed': 0,
            'hyperformula_compatible': 0,
            'python_required': 0,
            'exact_matches': 0,
            'within_tolerance': 0,
            'mismatches': 0,
            'errors': 0,
            'dropdown_cells': []
        }
    
    def _emit_progress(self, stage: str, percent: float, message: str):
        """Emit progress update via callback."""
        self.progress_callback(stage, percent, message)
        logger.info(f"Progress: {stage} ({percent:.1f}%) - {message}")
    
    def _topological_sort_formulas(self, cells_data: List[Dict]) -> List[List[Dict]]:
        """
        Sort formula cells in dependency order using topological sort.
        
        Returns batches of cells that can be evaluated in parallel.
        Cells in the same batch have no dependencies on each other.
        
        Algorithm:
            1. Build dependency graph (excluding circular cells)
            2. Use Kahn's algorithm for topological sort
            3. Group cells by evaluation level (batch)
        
        Args:
            cells_data: All cell data dictionaries (non-circular formulas only)
            
        Returns:
            List of batches, where each batch is a list of cells
            that can be evaluated in parallel
        """
        if not cells_data:
            return []
        
        # Build cell reference lookup
        cell_lookup = {}
        for cell in cells_data:
            cell_ref = f"{cell['sheet_name']}!{cell['cell']}"
            cell_lookup[cell_ref] = cell
        
        # Build in-degree map (how many dependencies each cell has)
        in_degree = {}
        for cell in cells_data:
            cell_ref = f"{cell['sheet_name']}!{cell['cell']}"
            in_degree[cell_ref] = 0
        
        # Count dependencies
        for cell in cells_data:
            cell_ref = f"{cell['sheet_name']}!{cell['cell']}"
            depends_on = cell.get('depends_on', [])
            
            # Only count dependencies that are in our cell set (non-circular formulas)
            for dep in depends_on:
                if dep in cell_lookup:
                    in_degree[cell_ref] += 1
        
        # Initialize queue with cells that have no dependencies
        queue = [cell for cell in cells_data if in_degree[f"{cell['sheet_name']}!{cell['cell']}"] == 0]
        batches = []
        
        # Process cells level by level (Kahn's algorithm)
        while queue:
            # All cells in current queue can be evaluated in parallel (same batch)
            current_batch = queue[:]
            batches.append(current_batch)
            queue = []
            
            # Process each cell in current batch
            for cell in current_batch:
                cell_ref = f"{cell['sheet_name']}!{cell['cell']}"
                
                # Find cells that depend on this cell
                for other_cell in cells_data:
                    other_ref = f"{other_cell['sheet_name']}!{other_cell['cell']}"
                    depends_on = other_cell.get('depends_on', [])
                    
                    if cell_ref in depends_on:
                        in_degree[other_ref] -= 1
                        
                        # If all dependencies are satisfied, add to next batch
                        if in_degree[other_ref] == 0 and other_cell not in current_batch:
                            queue.append(other_cell)
        
        logger.info(f"Topological sort: {len(batches)} evaluation batches")
        for i, batch in enumerate(batches[:5]):  # Log first 5 batches
            logger.debug(f"Batch {i}: {len(batch)} cells")
        
        return batches
    
    def _build_hyperformula_sheets(self, cells_data: List[Dict]) -> List[Dict]:
        """
        Build HyperFormula sheets data structure from parsed cells.
        
        Converts cell data to format expected by hyperformula_wrapper.js:
        {
            "name": "Sheet1",
            "cells": [
                {"row": 0, "col": 0, "formula": "=SUM(B1:B10)"},
                {"row": 0, "col": 1, "value": 5},
                {"row": 1, "col": 1, "value": 10}
            ]
        }
        
        Args:
            cells_data: All parsed cell data
            
        Returns:
            List of sheet dictionaries ready for HyperFormula
        """
        # Group cells by sheet
        sheets_dict = {}
        
        for cell in cells_data:
            sheet_name = cell['sheet_name']
            
            if sheet_name not in sheets_dict:
                sheets_dict[sheet_name] = {
                    'name': sheet_name,
                    'cells': []
                }
            
            # Convert cell address to coordinates
            try:
                row, col = self.parser.cell_to_coordinates(cell['cell'])
                
                # Add cell to sheet
                if cell.get('formula') and cell['cell_type'] != 'formula_text':
                    # Formula cell
                    sheets_dict[sheet_name]['cells'].append({
                        'row': row,
                        'col': col,
                        'formula': cell['formula']
                    })
                elif cell.get('raw_value') is not None:
                    # Numeric value cell (needed for formula dependencies)
                    sheets_dict[sheet_name]['cells'].append({
                        'row': row,
                        'col': col,
                        'value': float(cell['raw_value'])
                    })
                elif cell.get('raw_text') is not None:
                    # Text value cell (needed for IF conditions, string comparisons, etc.)
                    sheets_dict[sheet_name]['cells'].append({
                        'row': row,
                        'col': col,
                        'value': cell['raw_text']
                    })
                # Note: Skip text formulas (formula_text type)
                
            except Exception as e:
                logger.warning(f"Could not convert cell {sheet_name}!{cell['cell']}: {e}")
                continue
        
        # Convert to list
        sheets_list = list(sheets_dict.values())
        
        logger.info(f"Built HyperFormula sheets: {len(sheets_list)} sheets, "
                   f"{sum(len(s['cells']) for s in sheets_list)} cells")
        
        return sheets_list
    
    def _batch_evaluate_hyperformula(
        self,
        sheets_data: List[Dict],
        cells_to_evaluate: List[Dict],
        cache: Dict[str, float]
    ) -> Dict[str, Any]:
        """
        Batch evaluate formulas using HyperFormula.
        
        Args:
            sheets_data: HyperFormula sheets structure
            cells_to_evaluate: Cells to evaluate in this batch
            cache: Dictionary to cache evaluated values
            
        Returns:
            Dictionary mapping cell_ref to evaluated value/error
            
        Error Handling:
            - HyperFormula errors (#DIV/0!, #REF!, etc.) → return None
            - Evaluation failures → return None
            - Success → return numeric value
        """
        if not cells_to_evaluate:
            return {}
        
        # Build queries for this batch
        queries = []
        for cell in cells_to_evaluate:
            cell_ref = f"{cell['sheet_name']}!{cell['cell']}"
            
            # Skip if already cached
            if cell_ref in cache:
                continue
            
            try:
                row, col = self.parser.cell_to_coordinates(cell['cell'])
                queries.append({
                    'sheet': cell['sheet_name'],
                    'row': row,
                    'col': col,
                    'cell': cell_ref
                })
            except Exception as e:
                logger.error(f"Could not create query for {cell_ref}: {e}")
                continue
        
        if not queries:
            # All cells were cached
            return cache
        
        # Call HyperFormula
        logger.debug(f"Evaluating {len(queries)} formulas via HyperFormula")
        result = self.hf_evaluator.evaluate_batch(sheets_data, queries)
        
        if not result.get('success'):
            logger.error(f"HyperFormula batch evaluation failed: {result.get('error')}")
            # Return None for all cells in batch
            for query in queries:
                cache[query['cell']] = None
            return cache
        
        # Process results
        for res in result.get('results', []):
            cell_ref = res['cell']
            
            if res['type'] == 'number':
                # Successful numeric evaluation
                cache[cell_ref] = float(res['value'])
            elif res['type'] == 'error':
                # Excel error (#DIV/0!, #REF!, #VALUE!, etc.)
                error_value = res.get('value', 'ERROR')
                error_msg = res.get('error', '')
                
                # Log based on error severity
                if error_value in ['#DIV/0!', '#NULL!']:
                    logger.warning(f"Formula error in {cell_ref}: {error_value}")
                elif error_value in ['#REF!', '#NAME?']:
                    logger.error(f"Formula error in {cell_ref}: {error_value} - {error_msg}")
                else:
                    logger.info(f"Formula error in {cell_ref}: {error_value}")
                
                cache[cell_ref] = None
                self.stats['errors'] += 1
            else:
                # Empty or unexpected type
                logger.warning(f"Unexpected result type for {cell_ref}: {res['type']}")
                cache[cell_ref] = None
        
        return cache
    
    def compute_file_hash(self, file_path: str) -> str:
        """Compute SHA256 hash of file."""
        sha256_hash = hashlib.sha256()
        with open(file_path, "rb") as f:
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()
    
    def check_duplicate(self, file_hash: str) -> Optional[int]:
        """Check if file hash already exists in database."""
        existing = self.session.query(Model).filter_by(file_hash=file_hash).first()
        if existing:
            logger.info(f"File already imported as model ID {existing.id}")
            return existing.id
        return None
    
    def copy_to_models_dir(self, source_path: str, file_hash: str) -> str:
        """Copy Excel file to models directory."""
        models_path = Path(self.models_dir)
        models_path.mkdir(exist_ok=True)
        
        # Use hash as filename to avoid conflicts
        ext = Path(source_path).suffix
        dest_filename = f"{file_hash[:16]}{ext}"
        dest_path = models_path / dest_filename
        
        shutil.copy2(source_path, dest_path)
        logger.info(f"Copied file to {dest_path}")
        
        return str(dest_path)
    
    def parse_workbook(self, file_path: str) -> Dict:
        """
        Parse Excel workbook and extract all cell data.
        
        Returns dictionary with sheets and cells.
        """
        logger.info(f"Parsing workbook: {file_path}")
        
        # Load twice: once for formulas, once for computed values
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
        
        workbook_data = {
            'sheets': [],
            'cells': []
        }
        
        total_sheets = len(wb_formulas.sheetnames)
        
        for sheet_idx, sheet_name in enumerate(wb_formulas.sheetnames):
            ws_formulas = wb_formulas[sheet_name]
            ws_values = wb_values[sheet_name]
            
            # Calculate progress within parsing phase (10-30%)
            sheet_progress = 10 + (20 * (sheet_idx / total_sheets))
            self._emit_progress('parsing', sheet_progress, f"Processing sheet: {sheet_name}")
            
            logger.info(f"Processing sheet: {sheet_name}")
            
            sheet_info = {
                'name': sheet_name,
                'max_row': ws_formulas.max_row,
                'max_column': ws_formulas.max_column
            }
            workbook_data['sheets'].append(sheet_info)
            
            # Extract data validations (dropdowns)
            dropdown_cells = []
            if hasattr(ws_formulas, 'data_validations'):
                for dv in ws_formulas.data_validations.dataValidation:
                    if dv.type == 'list':
                        for cell_range in dv.cells:
                            dropdown_cells.append(f"{sheet_name}!{cell_range}")
            
            if dropdown_cells:
                self.stats['dropdown_cells'].extend(dropdown_cells)
            
            # Iterate through all cells in used range
            for row_idx, row in enumerate(ws_formulas.iter_rows(
                min_row=1, max_row=ws_formulas.max_row,
                min_col=1, max_col=ws_formulas.max_column
            ), 1):
                for cell in row:
                    if cell.value is None and not cell.data_type == 'f':
                        continue  # Skip empty cells
                    
                    # Get corresponding value cell
                    value_cell = ws_values.cell(row=cell.row, column=cell.column)
                    
                    cell_data = self.extract_cell_data(cell, value_cell, sheet_name, ws_formulas)
                    if cell_data:
                        workbook_data['cells'].append(cell_data)
                        self.stats['total_cells'] += 1
        
        logger.info(f"Parsed {len(workbook_data['sheets'])} sheets, "
                   f"{self.stats['total_cells']} cells")
        
        return workbook_data
    
    def extract_cell_data(self, cell_formula, cell_value, sheet_name: str, worksheet) -> Optional[Dict]:
        """
        Extract all data from a single cell.
        
        Args:
            cell_formula: Cell from workbook loaded with data_only=False (has formulas)
            cell_value: Cell from workbook loaded with data_only=True (has computed values)
            sheet_name: Name of the worksheet
            worksheet: Worksheet object for validation extraction
        """
        row_num = cell_formula.row
        col_letter = get_column_letter(cell_formula.column)
        cell_address = f"{col_letter}{row_num}"
        
        # Get formula
        formula = None
        if cell_formula.data_type == 'f':
            # Handle ArrayFormula objects from openpyxl
            if hasattr(cell_formula.value, 'text'):
                formula = cell_formula.value.text
            elif cell_formula.value:
                formula = str(cell_formula.value)
        
        # Classify cell type
        cell_type = 'value'
        if formula:
            if self.parser.is_text_formula(formula):
                cell_type = 'formula_text'
                self.stats['formula_text_cells'] += 1
            else:
                cell_type = 'formula'
                self.stats['formula_cells'] += 1
        else:
            self.stats['value_cells'] += 1
        
        # Get raw value from the data_only=True workbook (Excel's computed value)
        raw_value = None
        raw_text = None
        
        if cell_value.value is not None:
            try:
                # Try to convert to float
                if isinstance(cell_value.value, (int, float)):
                    raw_value = float(cell_value.value)
                elif isinstance(cell_value.value, str):
                    # Try to parse string as number
                    try:
                        raw_value = float(cell_value.value)
                    except ValueError:
                        # It's actually a text value - store in raw_text
                        raw_text = cell_value.value
                        logger.debug(f"Cell {cell_address} has text value: {raw_text}")
                else:
                    logger.debug(f"Cell {cell_address} has non-numeric value type: {type(cell_value.value)}")
            except (ValueError, TypeError) as e:
                logger.debug(f"Could not convert value for {cell_address}: {e}")
        
        # Infer data type from the computed value
        data_type = 'text'
        if isinstance(cell_value.value, (int, float)):
            data_type = 'number'
        elif isinstance(cell_value.value, bool):
            data_type = 'boolean'
        elif isinstance(cell_value.value, datetime):
            data_type = 'date'
        elif isinstance(cell_value.value, str):
            data_type = 'text'
        
        # Extract dependencies
        depends_on = []
        if formula:
            depends_on = self.parser.extract_dependencies(formula, sheet_name)
        
        # Check for validation
        has_validation = False
        validation_type = None
        validation_options = []
        
        cell_coord = cell_formula.coordinate
        if hasattr(worksheet, 'data_validations'):
            for dv in worksheet.data_validations.dataValidation:
                if cell_coord in dv.cells:
                    has_validation = True
                    validation_type = dv.type
                    if dv.formula1:
                        # Try to extract list values
                        try:
                            if dv.formula1.startswith('"'):
                                # Quoted list: "Option1,Option2,Option3"
                                options_str = dv.formula1.strip('"')
                                validation_options = [opt.strip() for opt in options_str.split(',')]
                            else:
                                # Range reference
                                validation_options = [dv.formula1]
                        except:
                            pass
                    break
        
        # Extract style information
        style = {}
        if cell_formula.font:
            style['font_size'] = cell_formula.font.size
            style['bold'] = cell_formula.font.bold
            style['italic'] = cell_formula.font.italic
        if cell_formula.border and cell_formula.border.left:
            style['border_style'] = cell_formula.border.left.style
        if cell_formula.fill and cell_formula.fill.start_color:
            # Safely extract RGB color
            try:
                if hasattr(cell_formula.fill.start_color, 'rgb'):
                    rgb = cell_formula.fill.start_color.rgb
                    # Check if it's actually a string (not an error message or validation object)
                    if isinstance(rgb, str) and not rgb.startswith('Values must be'):
                        style['bg_color'] = rgb
                    else:
                        style['bg_color'] = None
                else:
                    style['bg_color'] = None
            except Exception as e:
                logger.debug(f"Could not extract bg_color for {cell_address}: {e}")
                style['bg_color'] = None
        
        cell_data = {
            'sheet_name': sheet_name,
            'cell': cell_address,
            'row_num': row_num,
            'col_letter': col_letter,
            'cell_type': cell_type,
            'raw_value': raw_value,
            'raw_text': raw_text,
            'formula': formula,
            'data_type': data_type,
            'depends_on': depends_on,
            'has_validation': has_validation,
            'validation_type': validation_type,
            'validation_options': validation_options,
            'style': style
        }
        
        # For text value cells (non-formula), copy raw_text to calculated_text
        # This enables validation: we can compare calculated_text against raw_text
        if cell_type == 'value' and raw_text is not None:
            cell_data['calculated_text'] = raw_text
        
        return cell_data
    
    def import_file(self, file_path: str, model_name: str, validate: bool = False) -> Dict[str, Any]:
        """
        Main import workflow.
        
        Args:
            file_path: Path to Excel file
            model_name: Name for the model
            validate: Whether to run post-import validation
        
        Returns:
            Dictionary with import results:
            {
                'model_id': int,
                'stats': dict,
                'validation_results': dict or None,
                'errors': list
            }
        """
        logger.info(f"Starting import of {file_path} as '{model_name}'")
        errors = []
        
        try:
            # Step 1: Compute hash (5%)
            self._emit_progress('hashing', 5, 'Computing file hash...')
            file_hash = self.compute_file_hash(file_path)
            logger.info(f"File hash: {file_hash}")
            
            # Check for duplicate
            existing_id = self.check_duplicate(file_hash)
            if existing_id:
                logger.info("File already imported, skipping")
                return {
                    'model_id': existing_id,
                    'stats': {},
                    'validation_results': None,
                    'errors': ['File already imported'],
                    'duplicate': True
                }
            
            # Step 2: Copy to models directory (8%)
            self._emit_progress('copying', 8, 'Copying file to storage...')
            stored_path = self.copy_to_models_dir(file_path, file_hash)
            
            # Step 3: Parse workbook (10-30%)
            workbook_data = self.parse_workbook(file_path)
            
            # Step 4: Build dependency graph (30-35%)
            self._emit_progress('dependencies', 30, 'Building dependency graph...')
            logger.info("Building dependency graph...")
            for cell_data in workbook_data['cells']:
                cell_ref = f"{cell_data['sheet_name']}!{cell_data['cell']}"
                self.circular_detector.add_dependency(cell_ref, cell_data['depends_on'])
            
            circular_groups = self.circular_detector.detect_cycles()
            self.stats['circular_references'] = sum(len(group) for group in circular_groups)
            
            # Mark circular cells
            for cell_data in workbook_data['cells']:
                cell_ref = f"{cell_data['sheet_name']}!{cell_data['cell']}"
                cell_data['is_circular'] = self.circular_detector.is_circular(cell_ref)
            
            # Step 5: Create model record (37%)
            self._emit_progress('creating_model', 37, 'Creating model record...')
            workbook_meta = {
                'sheets': [s['name'] for s in workbook_data['sheets']],
                'sheet_count': len(workbook_data['sheets']),
                'total_cells': self.stats['total_cells'],
                'formula_cells': self.stats['formula_cells'],
                'dropdown_cells': self.stats['dropdown_cells']
            }
            
            model = Model(
                name=model_name,
                original_filename=Path(file_path).name,
                file_path=stored_path,
                file_hash=file_hash,
                workbook_metadata=workbook_meta,
                import_summary={}  # Will be updated after evaluation
            )
            
            self.session.add(model)
            self.session.flush()  # Get model.id
            
            logger.info(f"Created model record with ID {model.id}")
            
            # Step 6: Evaluate formulas (40-80%)
            self._emit_progress('evaluation', 40, 'Evaluating formulas...')
            logger.info("Evaluating formulas...")
            self.evaluate_formulas(workbook_data['cells'])
            
            # Step 7: Bulk insert cells (80-95%)
            self._emit_progress('insertion', 80, 'Inserting cells into database...')
            logger.info("Inserting cells into database...")
            self.bulk_insert_cells(model.id, workbook_data['cells'])
            
            # Step 8: Update import summary (96%)
            self._emit_progress('finalizing', 96, 'Finalizing import...')
            model.import_summary = {
                **self.stats,
                'tolerance_used': self.tolerance,
                'import_timestamp': datetime.utcnow().isoformat()
            }
            
            self.session.commit()
            
            logger.info(f"Import completed successfully. Model ID: {model.id}")
            
            # Step 9: Optional validation (97-99%)
            validation_result = None
            if validate:
                self._emit_progress('validation', 97, 'Running post-import validation...')
                logger.info("Running post-import validation...")
                from services.validation_service import ValidationService
                validator = ValidationService(self.session, self.progress_callback)
                validation_result = validator.validate_model(model.id)
                
                # Update model with validation results
                model.import_summary['post_import_validation'] = validation_result
                self.session.commit()
                
                logger.info(f"Validation completed: {validation_result}")
            
            # Step 10: Complete (100%)
            self._emit_progress('complete', 100, 'Import complete')
            
            return {
                'model_id': model.id,
                'stats': self.stats,
                'validation_results': validation_result,
                'errors': errors,
                'duplicate': False
            }
            
        except Exception as e:
            logger.error(f"Import failed: {e}", exc_info=True)
            self.session.rollback()
            errors.append(str(e))
            return {
                'model_id': None,
                'stats': self.stats,
                'validation_results': None,
                'errors': errors,
                'duplicate': False
            }
    
    def evaluate_formulas(self, cells_data: List[Dict]):
        """
        Evaluate all formulas using HyperFormula with proper dependency ordering.
        
        This implements the core evaluation pipeline with progress tracking,
        topological sorting, and batch evaluation for efficiency.
        """
        total_formulas = self.stats['formula_cells'] + self.stats['formula_text_cells']
        logger.info(f"Evaluating {total_formulas} formula cells...")
        
        # Build HyperFormula data structure (include ALL cells for context)
        self._emit_progress('evaluation', 40, 'Building formula context...')
        sheets_data = self._build_hyperformula_sheets(cells_data)
        
        # Initialize evaluation cache
        cache = {}
        
        # Build lookup for quick access
        cell_lookup = {}
        for cell in cells_data:
            cell_ref = f"{cell['sheet_name']}!{cell['cell']}"
            cell_lookup[cell_ref] = cell
        
        # Separate circular from non-circular
        circular_cells = []
        non_circular_cells = []
        
        for cell in cells_data:
            if cell.get('formula'):
                if cell.get('is_circular'):
                    circular_cells.append(cell)
                else:
                    non_circular_cells.append(cell)
        
        logger.info(f"Non-circular formulas: {len(non_circular_cells)}, "
                   f"Circular formulas: {len(circular_cells)}")
        
        # Topological sort for evaluation order (42%)
        self._emit_progress('evaluation', 42, 'Sorting formulas by dependencies...')
        evaluation_batches = self._topological_sort_formulas(non_circular_cells)
        
        # Evaluate non-circular formulas in dependency order (45-70%)
        total_batches = len(evaluation_batches)
        evaluated = 0
        
        for batch_idx, batch in enumerate(evaluation_batches):
            # Progress tracking
            progress = 45 + (25 * (batch_idx / max(total_batches, 1)))
            self._emit_progress('evaluation', progress,
                              f"Evaluating batch {batch_idx+1}/{total_batches} ({len(batch)} formulas)")
            
            # Evaluate this batch
            self._evaluate_batch(batch, sheets_data, cache, cell_lookup)
            evaluated += len(batch)
        
        # Evaluate circular formulas with iterative solver (70-80%)
        if circular_cells:
            self._emit_progress('evaluation', 70, 'Solving circular references...')
            self._evaluate_circular_cells_hyperformula(
                circular_cells,
                sheets_data,
                cell_lookup,
                cache
            )
        
        logger.info("Formula evaluation complete")
    
    def _evaluate_batch(
        self,
        batch: List[Dict],
        sheets_data: List[Dict],
        cache: Dict[str, float],
        cell_lookup: Dict
    ):
        """
        Evaluate a batch of cells using HyperFormula.
        
        Cells in a batch can be evaluated in parallel since they
        have no dependencies on each other.
        """
        # Separate numeric formulas from text formulas
        numeric_formulas = []
        text_formulas = []
        
        for cell in batch:
            if cell['cell_type'] == 'formula_text':
                text_formulas.append(cell)
            else:
                numeric_formulas.append(cell)
        
        # Evaluate text formulas (simple Python evaluation)
        for cell in text_formulas:
            formula = cell.get('formula', '')
            
            cell['calculation_engine'] = 'hyperformula'
            cell['converted_formula'] = formula
            
            try:
                result_text = self.parser.evaluate_text_formula(formula)
                cell['calculated_text'] = result_text
                cell['calculated_value'] = None
                
                # Compare with raw_text if available
                if result_text is not None and cell.get('raw_text') is not None:
                    if result_text == cell['raw_text']:
                        self.stats['exact_matches'] += 1
                    else:
                        cell['has_mismatch'] = True
                        cell['mismatch_diff'] = float(abs(len(result_text) - len(cell['raw_text'])))
                        self.stats['mismatches'] += 1
            except Exception as e:
                logger.error(f"Text formula evaluation failed for {cell['sheet_name']}!{cell['cell']}: {e}")
                cell['calculated_text'] = None
                self.stats['errors'] += 1
        
        # Batch evaluate numeric formulas through HyperFormula
        if numeric_formulas:
            self._batch_evaluate_hyperformula(sheets_data, numeric_formulas, cache)
            
            # Apply results to cells
            for cell in numeric_formulas:
                cell_ref = f"{cell['sheet_name']}!{cell['cell']}"
                result_value = cache.get(cell_ref)
                
                # Classify engine type
                formula = cell.get('formula', '')
                if self.parser.is_custom_function(formula):
                    cell['calculation_engine'] = 'custom'
                    cell['converted_formula'] = self.parser.convert_for_custom(formula)
                else:
                    cell['calculation_engine'] = 'hyperformula'
                    cell['converted_formula'] = formula
                
                cell['calculated_value'] = result_value
                cell['calculated_text'] = None
                
                # Compare with raw_value if available
                if result_value is not None and cell.get('raw_value') is not None:
                    diff = abs(float(result_value) - float(cell['raw_value']))
                    if diff > self.tolerance:
                        cell['has_mismatch'] = True
                        cell['mismatch_diff'] = float(diff)
                        self.stats['mismatches'] += 1
                    elif diff < 1e-10:
                        self.stats['exact_matches'] += 1
                    else:
                        self.stats['within_tolerance'] += 1
                
                # Track stats
                if cell['calculation_engine'] == 'hyperformula':
                    self.stats['hyperformula_compatible'] += 1
                elif cell['calculation_engine'] == 'custom':
                    self.stats['python_required'] += 1
    
    def _evaluate_numeric_formula(
        self,
        cell: Dict,
        cell_lookup: Dict,
        sheets_data: List[Dict],
        cache: Dict[str, float]
    ) -> Optional[float]:
        """
        Evaluate numeric formula using HyperFormula.
        
        This replaces the placeholder implementation that used raw_value.
        
        Args:
            cell: Cell data dictionary
            cell_lookup: All cells indexed by reference
            sheets_data: HyperFormula sheets structure
            cache: Evaluation cache
            
        Returns:
            Evaluated numeric value or None on error
        """
        formula = cell.get('formula', '')
        cell_ref = f"{cell['sheet_name']}!{cell['cell']}"
        
        # Check cache first
        if cell_ref in cache:
            return cache[cell_ref]
        
        # Simple constant formulas (fast path)
        if re.match(r'^=\d+(\.\d+)?$', formula):
            result = float(formula[1:])
            cache[cell_ref] = result
            return result
        
        # Evaluate through HyperFormula
        try:
            row, col = self.parser.cell_to_coordinates(cell['cell'])
            
            result = self.hf_evaluator.evaluate_batch(
                sheets_data=sheets_data,
                queries=[{
                    'sheet': cell['sheet_name'],
                    'row': row,
                    'col': col,
                    'cell': cell_ref
                }]
            )
            
            if result.get('success') and result['results']:
                res = result['results'][0]
                
                # Handle different result types
                if res['type'] == 'number':
                    value = float(res['value'])
                    cache[cell_ref] = value
                    return value
                elif res['type'] == 'error':
                    # Excel errors: #DIV/0!, #REF!, #VALUE!, etc.
                    error_value = res.get('value', 'ERROR')
                    logger.warning(f"Formula error for {cell_ref}: {error_value}")
                    return None
                else:
                    # Empty or unexpected type
                    return None
            else:
                logger.error(f"HyperFormula evaluation failed for {cell_ref}: {result.get('error')}")
                return None
                
        except Exception as e:
            logger.error(f"Error evaluating {cell_ref}: {e}")
            return None
    
    def _evaluate_circular_cells_hyperformula(
        self,
        circular_cells: List[Dict],
        sheets_data: List[Dict],
        cell_lookup: Dict,
        cache: Dict
    ):
        """
        Evaluate circular formulas using HyperFormula's built-in circular reference handling.
        
        HyperFormula has built-in iterative calculation for circular references.
        We just need to evaluate them normally - HyperFormula handles the iteration internally.
        """
        logger.info(f"Evaluating {len(circular_cells)} circular formulas with HyperFormula...")
        
        # Build queries for all circular cells
        queries = []
        for cell in circular_cells:
            try:
                row, col = self.parser.cell_to_coordinates(cell['cell'])
                cell_ref = f"{cell['sheet_name']}!{cell['cell']}"
                queries.append({
                    'sheet': cell['sheet_name'],
                    'row': row,
                    'col': col,
                    'cell': cell_ref
                })
            except Exception as e:
                logger.error(f"Could not create query for circular cell: {e}")
                continue
        
        # Evaluate through HyperFormula (it handles circular references internally)
        logger.debug(f"Evaluating {len(queries)} circular formulas via HyperFormula")
        result = self.hf_evaluator.evaluate_batch(sheets_data, queries)
        
        if not result.get('success'):
            logger.error(f"HyperFormula circular evaluation failed: {result.get('error')}")
            # Set all to custom/None on failure
            for cell in circular_cells:
                cell['calculation_engine'] = 'custom'
                cell['converted_formula'] = cell.get('formula')
                cell['calculated_value'] = None
                cell['calculated_text'] = None
                cell['has_mismatch'] = False  # Not a mismatch - HyperFormula limitation
                cell['mismatch_diff'] = None
                self.stats['circular_failed'] += 1
                self.stats['python_required'] += 1
            return
        
        # Process results
        values = {}
        for res in result['results']:
            cell_ref = res['cell']
            
            if res['type'] == 'number':
                values[cell_ref] = float(res['value'])
            elif res['type'] == 'error':
                # Circular references return #CYCLE! errors
                error_val = res.get('value', 'ERROR')
                logger.info(f"Circular cell {cell_ref} has error: {error_val}")
                values[cell_ref] = None
            else:
                logger.warning(f"Unexpected type for circular cell {cell_ref}: {res['type']}")
                values[cell_ref] = None
        
        # Apply results to cells
        for cell in circular_cells:
            cell_ref = f"{cell['sheet_name']}!{cell['cell']}"
            result_value = values.get(cell_ref)
            
            # Circular cells that HyperFormula can't evaluate are marked as 'custom'
            if result_value is None:
                cell['calculation_engine'] = 'custom'
                cell['converted_formula'] = cell.get('formula')
                cell['calculated_value'] = None
                cell['calculated_text'] = None
                cell['has_mismatch'] = False  # Not a mismatch - just unsupported
                cell['mismatch_diff'] = None
                self.stats['circular_failed'] += 1
                self.stats['python_required'] += 1
                logger.debug(f"Circular cell {cell_ref} marked as 'custom' (HyperFormula limitation)")
            else:
                # Successfully evaluated circular cell
                cell['calculation_engine'] = 'hyperformula'
                cell['converted_formula'] = cell.get('formula')
                cell['calculated_value'] = result_value
                cell['calculated_text'] = None
                
                # Compare with raw_value for validation
                if cell.get('raw_value') is not None:
                    diff = abs(float(result_value) - float(cell['raw_value']))
                    if diff > self.tolerance:
                        cell['has_mismatch'] = True
                        cell['mismatch_diff'] = float(diff)
                        self.stats['mismatches'] += 1
                    elif diff < 1e-10:
                        self.stats['exact_matches'] += 1
                    else:
                        self.stats['within_tolerance'] += 1
                
                # Track stats and cache
                self.stats['circular_converged'] += 1
                self.stats['hyperformula_compatible'] += 1
                cache[cell_ref] = result_value
        
        logger.info(f"Circular evaluation complete: {self.stats['circular_converged']} converged, "
                   f"{self.stats['circular_failed']} failed")
    
    
    def bulk_insert_cells(self, model_id: int, cells_data: List[Dict]):
        """Bulk insert cells in batches."""
        BATCH_SIZE = 1000
        total_cells = len(cells_data)
        
        for i in range(0, total_cells, BATCH_SIZE):
            batch = cells_data[i:i + BATCH_SIZE]
            cell_objects = []
            
            # Update progress
            progress = 80 + (15 * (i / max(total_cells, 1)))
            self._emit_progress('insertion', progress, 
                              f"Inserting cells {i}/{total_cells}")
            
            for cell_data in batch:
                cell_obj = Cell(
                    model_id=model_id,
                    **{k: v for k, v in cell_data.items() if k in [
                        'sheet_name', 'cell', 'row_num', 'col_letter', 'cell_type',
                        'raw_value', 'raw_text', 'formula', 'data_type', 'depends_on', 'is_circular',
                        'has_validation', 'validation_type', 'validation_options',
                        'style', 'calculation_engine', 'converted_formula',
                        'calculated_value', 'calculated_text', 'has_mismatch', 'mismatch_diff'
                    ]}
                )
                cell_objects.append(cell_obj)
            
            self.session.bulk_save_objects(cell_objects)
            self.session.flush()
            
            logger.debug(f"Inserted batch {i//BATCH_SIZE + 1} ({len(batch)} cells)")
        
        logger.info(f"Inserted {len(cells_data)} cells")